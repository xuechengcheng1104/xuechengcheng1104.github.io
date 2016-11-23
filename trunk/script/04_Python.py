
#------------------------------------------------------------------------------
>>    Windows 安装库openpyxl
Step 1: 下载pip-8.1.2.tar.gz，解压
https://pypi.python.org/pypi/pip#downloads
Step 2: 安装pip
python setup.py install
Step 3: 安装openpyxl
pip install openpyxl
#-------------------------------------------------
#    python.py脚本文件格式
#-------------------------------------------------
UNIX and UTF-8
#-------------------------------------------------
#    openpyxl class
#-------------------------------------------------
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
dest_filename = 'output.xlsx'
#构造类
wb = Workbook()
wb2 = load_workbook('output.xlsx')
#新建子类worksheet
ws = wb.active
ws1 = wb.create_sheet()
ws2 = wb.create_sheet(0)
ws21 = wb.create_sheet(title="Pi")
#子类worksheet属性修改
ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA"
#获取Workbook子类worksheet实例
ws3 = wb["New Title"]
ws4 = wb.get_sheet_by_name("New Title")
#输出worksheet名称
print(wb.get_sheet_names())
for sheet in wb:
    print(sheet.title)
#单元格赋值
c = ws['A4']
d = ws.cell(row = 4, column = 2)
ws['A1'] = 42
ws.append([1, 2, 3])
ws['A3'] = datetime.datetime.now()
ws['A4'] = datetime.datetime(2010, 7, 21)
ws["A5"] = "=SUM(1, 1)"
for i in range(1,101):
    for j in range(1,101):
        ws.cell(row = i, column = j).value = i*100+j
#单元格值输出
for row in ws['A1:C2']:
    for cell in row:
        print (cell)
        print (cell.value)
print (ws['A1'].number_format)
print(ws3['AA10'].value)
#合并单元格
ws.merge_cells('A1:B1')
ws.unmerge_cells('A1:B1')
ws.merge_cells(start_row=2,start_column=2,end_row=4,end_column=4)
#输出类实例到文件
wb.save("output.xlsx")
wb.save(filename = dest_filename)
#-------------------------------------------------
#    import os
#-------------------------------------------------
os.getcwd()
os.name
os.remove()
os.removedirs()
os.system()
os.mkdir()
os.chdir()
os.listdir()
#-------------------------------------------------
#    读入文件内容，整理成列表，输入文件
#-------------------------------------------------
import re
with open('output.q', 'r+') as f:
	text = f.read()
	entries = re.split("\n+", text)
	for entry in entries:
		v1_list = re.split(":? ", entry, 4)
		v2_string = str(v1_list)
		v2_string = v2_string.replace("[","")
		v2_string = "\n"+v2_string.replace("]","")
		f.write(v2_string)
f.close()
#-------------------------------------------------
#    替换参数文件中的参数值
#-------------------------------------------------
#!/user/bin/python
import sys
import re
import datetime

params=sys.argv
fobj=open(params[1], 'r')

def defineFileType(filename):
	array=filename.split('.')
	return array[len(array)-1]
def getJobStartDate():
	return params[2]
def jobStartChange(job_start):
	time=job_start[job_start.index('T'):]
	return 'job_start='+getJobStartDate()+time
def getJobEndDate():
	return params[3]
def jobEndChange(job_end):
	time=job_end[job_end.index('T'):]
	return 'job_end='+getJobEndDate()+time
def replacePropValue(input_str):
	if input_str=='':
		return input_str
	str_key=input_str[0:input_str.index('=')]
	if str_key=='job_start':
		return jobStartChange(input_str)
	if str_key=='job_end':
		return jobEndChange(input_str)
	return input_str

output=""
try:
	if defineFileType(params[1])=='properties':
		for line in fobj:
			if line.startswith('#'): #return true if this line start with '#'
				output+=line.strip() #删除首尾指定字符（默认为空格）
				output+='\n'
				continue
			tmp=line.strip()
			flag=replacePropValue(line.strip()) #this the the key
			if flag != line.strip():
				tmp=flag
			output+=tmp
			output+='\n'
finally:
	fobj.close()

f=open(params[1],'w') #写文件，'w+'追加文件
f.write(output)
f.close()